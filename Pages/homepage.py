import openpyxl
from POMTest.Locators.locators import Locators
class HomePage():


    def __init__(self, driver):
        self.driver=driver
        self.audio_clk = Locators.audio_clk
        self.home_theatre_clk = Locators.home_theatre_clk
        self.product_clk = Locators.product_clk
        self.add_to_cart = Locators.add_to_cart
        self.checkout = Locators.checkout
        self.checkout_guest=Locators.checkout_guest
        self.company = Locators.company
        self.first_name = Locators.first_name
        self.last_name = Locators.last_name
        self.address_1 = Locators.address_1
        self.address_2 = Locators.address_2
        self.city = Locators.city
        self.zip_code = Locators.zip_code
        self.select_country = Locators.select_country
        self.country = Locators.country
        self.select_state = Locators.select_state
        self.state = Locators.state
        self.email = Locators.email
        self.phone_number = Locators.phone_number
        self.next_button = Locators.next_button
        self.ship_button = Locators.ship_button
        self.next_button1 = Locators.next_button1
        self.next_button2 = Locators.next_button2


    def click_audio(self):
        self.driver.find_element_by_xpath(self.audio_clk).click()

    def click_home_theatre(self):
        self.driver.find_element_by_xpath(self.home_theatre_clk).click()

    def click_product(self):
        self.driver.find_element_by_xpath(self.product_clk).click()

    def click_add_cart(self):
        self.driver.find_element_by_xpath(self.add_to_cart).click()

    def click_checkout_btn(self):
        self.driver.find_element_by_class_name(self.checkout).click()

    def click_checkout_guest(self):
        self.driver.find_element_by_css_selector(self.checkout_guest).click()

    def click_billing_address(self):
        workbook=openpyxl.load_workbook("D:\Testing\ReadExcel\Sathya.xlsx")
        sheet=workbook['Details']
        #self.driver.find_element_by_css_selector(self.company).send_keys(sheet['B1'].value)
        #self.driver.find_element_by_css_selector(self.company).send_keys(sheet.cell(1,2).value)
        self.driver.find_element_by_css_selector(self.company).send_keys(sheet.cell(column=2,row=1).value)
        self.driver.find_element_by_css_selector(self.first_name).send_keys(sheet['B2'].value)
        self.driver.find_element_by_css_selector(self.last_name).send_keys(sheet['B3'].value)
        self.driver.find_element_by_css_selector(self.address_1).send_keys(sheet['B4'].value)
        self.driver.find_element_by_css_selector(self.address_2).send_keys(sheet['B5'].value)
        self.driver.find_element_by_css_selector(self.city).send_keys(sheet['B6'].value)
        self.driver.find_element_by_css_selector(self.zip_code).send_keys(sheet['B7'].value)
        self.driver.find_element_by_css_selector(self.select_country).click()
        self.driver.find_element_by_xpath(self.country).click()
        self.driver.find_element_by_css_selector(self.select_state).click()
        self.driver.find_element_by_xpath(self.state).click()
        self.driver.find_element_by_css_selector(self.email).send_keys(sheet['B8'].value)
        self.driver.find_element_by_css_selector(self.phone_number).send_keys(sheet['B9'].value)
        self.driver.find_element_by_xpath(self.next_button).click()

    def click_ship_address(self):
        self.driver.find_element_by_xpath(self.ship_button).click()

    def click_next_button1(self):
        self.driver.find_element_by_xpath(self.next_button1).click()

    def click_next_button2(self):
        self.driver.find_element_by_xpath(self.next_button2).click()

